import io
import json
import os
import smtplib
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from typing import List, Optional

import anthropic
import openpyxl
import pdfplumber
import yaml
from fastapi import APIRouter, File, HTTPException, UploadFile
from pydantic import BaseModel

router = APIRouter()

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
PERSONAL_DATA_PATH = os.path.join(ROOT, "personal_data.yml")
RESUME_CONTEXT_PATH = os.path.join(ROOT, "resume_context.yml")
RESUME_PDF_PATH = os.path.join(ROOT, "resume_cached.pdf")


def load_personal_data():
    if not os.path.exists(PERSONAL_DATA_PATH):
        raise HTTPException(
            status_code=404,
            detail="personal_data.yml not found. Copy personal_data.yml.example and fill in your details.",
        )
    with open(PERSONAL_DATA_PATH) as f:
        return yaml.safe_load(f)


def load_resume_context():
    if not os.path.exists(RESUME_CONTEXT_PATH):
        return None
    with open(RESUME_CONTEXT_PATH) as f:
        return yaml.safe_load(f)


class Contact(BaseModel):
    company_name: str
    role: str
    manager_name: Optional[str] = None
    manager_email: str
    company_description: Optional[str] = None
    job_url: Optional[str] = None


class EmailDraft(BaseModel):
    id: str
    to_name: Optional[str] = None
    to_email: str
    company_name: str
    role: str
    subject: str
    body: str


class SendRequest(BaseModel):
    drafts: List[EmailDraft]
    attach_resume: bool = True


@router.post("/parse-resume")
async def parse_resume(file: UploadFile = File(...)):
    content = await file.read()

    with open(RESUME_PDF_PATH, "wb") as f:
        f.write(content)

    text = ""
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page in pdf.pages:
            text += page.extract_text() or ""

    client = anthropic.Anthropic()

    response = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=1500,
        messages=[
            {
                "role": "user",
                "content": f"""Extract structured information from this resume into YAML. Be concise. Include:
- name
- email
- current_status (e.g. "Final year B.Tech student", "working professional")
- education (institution, degree, cgpa, year)
- experience (list with: company, role, duration, highlights as 2-3 bullet strings)
- skills (grouped by category as lists)
- achievements (hackathon wins, notable projects as brief strings)
- portfolio_links (linkedin, github, portfolio url)

Resume:
{text}

Return ONLY valid YAML, no markdown fences, no extra text.""",
            }
        ],
    )

    yaml_text = response.content[0].text.strip()
    if yaml_text.startswith("```"):
        yaml_text = yaml_text.split("\n", 1)[1].rsplit("```", 1)[0].strip()

    try:
        context = yaml.safe_load(yaml_text)
    except yaml.YAMLError as e:
        raise HTTPException(status_code=500, detail=f"Failed to parse resume: {e}")

    with open(RESUME_CONTEXT_PATH, "w") as f:
        yaml.dump(context, f, default_flow_style=False, allow_unicode=True)

    return {"status": "ok", "context": context}


@router.get("/resume-context")
def get_resume_context():
    ctx = load_resume_context()
    return {"context": ctx, "cached": ctx is not None}


@router.post("/parse-contacts")
async def parse_contacts(file: UploadFile = File(...)):
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h).strip().lower().replace(" ", "_") if h else "" for h in header_row]

    col = {}
    for i, h in enumerate(headers):
        if "company" in h and "company_name" not in col:
            col["company_name"] = i
        elif any(k in h for k in ("role", "position", "title")) and "role" not in col:
            col["role"] = i
        elif "manager" in h and "name" in h and "manager_name" not in col:
            col["manager_name"] = i
        elif "email" in h and "manager_email" not in col:
            col["manager_email"] = i
        elif any(k in h for k in ("description", "about")) and "company_description" not in col:
            col["company_description"] = i
        elif any(k in h for k in ("url", "link", "job_url")) and "job_url" not in col:
            col["job_url"] = i

    missing = [k for k in ("company_name", "manager_email", "role") if k not in col]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required columns: {missing}. Required: Company Name, Role, Manager Email",
        )

    def cell(row, key):
        idx = col.get(key)
        if idx is None or idx >= len(row):
            return None
        val = row[idx]
        return str(val).strip() if val is not None else None

    contacts = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        company = cell(row, "company_name")
        email = cell(row, "manager_email")
        if not company or not email:
            continue
        contacts.append(
            {
                "company_name": company,
                "role": cell(row, "role") or "Software Engineer",
                "manager_name": cell(row, "manager_name"),
                "manager_email": email,
                "company_description": cell(row, "company_description"),
                "job_url": cell(row, "job_url"),
            }
        )

    return {"contacts": contacts, "count": len(contacts)}


@router.post("/generate")
async def generate_emails(contacts: List[Contact]):
    resume_ctx = load_resume_context()
    if not resume_ctx:
        raise HTTPException(
            status_code=400,
            detail="No cached resume. Upload a resume first.",
        )

    personal = load_personal_data()
    sender = personal.get("sender", {})
    sender_name = sender.get("name", "")

    client = anthropic.Anthropic()

    system_prompt = f"""You write concise, effective cold emails for job applications on behalf of {sender_name}.

Candidate profile:
{yaml.dump(resume_ctx, default_flow_style=False, allow_unicode=True)}

Contact details to include in every email:
- Email: {sender.get("email", "")}
- LinkedIn: {sender.get("linkedin", "")}
- GitHub: {sender.get("github", "")}
{f'- Portfolio: {sender.get("portfolio", "")}' if sender.get("portfolio") else ""}

Rules:
- Body: 150-200 words max
- No generic opener like "I hope this finds you well"
- No em dashes
- Reference 1-2 specific experiences or projects most relevant to the role
- End with a single, low-pressure ask (15-minute call or reply)
- Subject line: specific, role-focused, not clickbait

Respond with ONLY a raw JSON object: {{"subject": "...", "body": "..."}}
The body should use plain text with \\n for newlines."""

    drafts = []

    for i, contact in enumerate(contacts):
        user_msg = f"""Write a cold email for this target:

Company: {contact.company_name}
Role: {contact.role}
Hiring Manager: {contact.manager_name or "Hiring Manager"}
{f"About the company: {contact.company_description}" if contact.company_description else ""}
{f"Job posting: {contact.job_url}" if contact.job_url else ""}

Address it to {contact.manager_name or "the hiring team"}."""

        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=600,
            system=[
                {
                    "type": "text",
                    "text": system_prompt,
                    "cache_control": {"type": "ephemeral"},
                }
            ],
            messages=[{"role": "user", "content": user_msg}],
        )

        raw = response.content[0].text.strip()
        if raw.startswith("```"):
            raw = raw.split("\n", 1)[1].rsplit("```", 1)[0].strip()

        try:
            parsed = json.loads(raw)
            subject = parsed["subject"]
            body = parsed["body"]
        except Exception:
            subject = f"{contact.role} - {sender_name}"
            body = raw

        drafts.append(
            {
                "id": f"draft_{i}",
                "to_name": contact.manager_name,
                "to_email": contact.manager_email,
                "company_name": contact.company_name,
                "role": contact.role,
                "subject": subject,
                "body": body,
            }
        )

    return {"drafts": drafts}


@router.post("/send")
async def send_emails(req: SendRequest):
    personal = load_personal_data()
    sender = personal.get("sender", {})

    gmail_email = sender.get("email")
    gmail_password = sender.get("gmail_app_password")
    sender_name = sender.get("name", gmail_email)

    if not gmail_email or not gmail_password:
        raise HTTPException(
            status_code=400,
            detail="Gmail credentials missing in personal_data.yml",
        )

    results = []

    try:
        smtp = smtplib.SMTP_SSL("smtp.gmail.com", 465)
        smtp.login(gmail_email, gmail_password)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Gmail login failed: {e}")

    for draft in req.drafts:
        try:
            msg = MIMEMultipart()
            msg["From"] = f"{sender_name} <{gmail_email}>"
            msg["To"] = draft.to_email
            msg["Subject"] = draft.subject

            msg.attach(MIMEText(draft.body, "plain"))

            if req.attach_resume and os.path.exists(RESUME_PDF_PATH):
                with open(RESUME_PDF_PATH, "rb") as f:
                    attach = MIMEApplication(f.read(), _subtype="pdf")
                    fname = f"{sender_name.replace(' ', '_')}_Resume.pdf"
                    attach.add_header("Content-Disposition", "attachment", filename=fname)
                    msg.attach(attach)

            smtp.sendmail(gmail_email, draft.to_email, msg.as_string())
            results.append({"id": draft.id, "to": draft.to_email, "status": "sent"})
        except Exception as e:
            results.append({"id": draft.id, "to": draft.to_email, "status": "failed", "error": str(e)})

    smtp.quit()
    return {"results": results}
